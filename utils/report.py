"""
Report Generation Module
Excel report creation for Handover data
"""

import io
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None

CHANNELS = ["SPX", "GHN"]
THIN = Side(style="thin", color="CCCCCC")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")


def now_vn():
    """Get current time in Vietnam timezone"""
    if ZoneInfo:
        return datetime.now(ZoneInfo("Asia/Ho_Chi_Minh"))
    return datetime.utcnow()


def _autosize_columns(ws):
    """Auto-size worksheet columns based on content"""
    maxw = {}
    for row in ws.iter_rows(values_only=True):
        for c_idx, val in enumerate(row, start=1):
            ln = len(str(val)) if val is not None else 0
            maxw[c_idx] = max(maxw.get(c_idx, 0), ln)

    for c_idx, w in maxw.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(w + 2, 10), 50)


def _style_all_center(ws):
    """Apply centered styling to worksheet"""
    # Header
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = CENTER
        cell.border = BORDER_ALL

    # Body
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = CENTER
            cell.border = BORDER_ALL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22
    _autosize_columns(ws)


def _parse_scan_ts(ev: dict) -> int:
    """Parse scan timestamp from event data"""
    ts = ev.get("ts")
    if isinstance(ts, (int, float)) and ts > 0:
        return int(ts)

    s = ev.get("time_vn") or ev.get("time") or ""
    if not s:
        return 0

    try:
        return int(datetime.strptime(s, "%Y-%m-%d %H:%M:%S").timestamp())
    except ValueError:
        try:
            return int(datetime.strptime(s, "%d-%m-%Y %H:%M:%S").timestamp())
        except ValueError:
            return 0


def build_report_message(date_str, snap):
    """Build report message text with statistics"""
    count_cancel = 0
    per_ch_non_cancel = {ch: 0 for ch in CHANNELS}

    for ch in CHANNELS:
        data_ch = (snap.get(ch) or {})
        for _, ev in data_ch.items():
            if (ev or {}).get("is_cancelled") == "Yes":
                count_cancel += 1
            else:
                per_ch_non_cancel[ch] += 1

    total_non_cancel = sum(per_ch_non_cancel.values())
    t = now_vn()
    current_time = t.strftime("%H:%M:%S")

    msg = (
        f"**[- VNDB/L -] REPORT HANDOVER:**\n"
        f"Ngày {date_str} đã bàn giao tổng {total_non_cancel} đơn:\n"
        f"- SPX: {per_ch_non_cancel['SPX']} đơn\n"
        f"- GHN: {per_ch_non_cancel['GHN']} đơn\n"
        f"- Total Cancel: {count_cancel} đơn\n\n"
        f"***Updated lúc {current_time}***"
    )

    filename = f"DataHandover_{date_str} - {total_non_cancel}.xlsx"
    return msg, filename, per_ch_non_cancel, count_cancel, total_non_cancel


def create_excel_report(snap, filename):
    """Create Excel workbook with handover data"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for ch in CHANNELS:
        ws = wb.create_sheet(ch)
        ws.append(["STT", "Scan Time", "LM Tracking", "Người Bàn Giao"])

        data_ch = (snap.get(ch) or {})
        sorted_items = sorted(
            data_ch.items(),
            key=lambda kv: _parse_scan_ts(kv[1]),
            reverse=True
        )

        for i, (order_id, ev) in enumerate(sorted_items, start=1):
            ws.append([
                i,
                ev.get("time_vn") or ev.get("time") or "",
                order_id,
                ev.get("user", "")
            ])

        _style_all_center(ws)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
