"""
Excel Utilities
"""

from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

THIN = Side(style="thin", color="CCCCCC")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")


def autosize_columns(ws):
    """Auto-size worksheet columns"""
    maxw = {}
    for row in ws.iter_rows(values_only=True):
        for c_idx, val in enumerate(row, start=1):
            ln = len(str(val)) if val is not None else 0
            maxw[c_idx] = max(maxw.get(c_idx, 0), ln)

    for c_idx, w in maxw.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(w + 2, 10), 50)


def style_all_center(ws):
    """Apply centered styling to worksheet"""
    # Header
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = CENTER
        cell.border = BORDER_ALL

    # Body
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = CENTER
            cell.border = BORDER_ALL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22
    autosize_columns(ws)


def parse_scan_ts(ev: dict) -> int:
    """Parse timestamp from event"""
    ts = ev.get("ts")
    if isinstance(ts, (int, float)) and ts > 0:
        return int(ts)

    s = ev.get("time_vn") or ev.get("time") or ""
    if not s:
        return 0

    try:
        return int(datetime.strptime(s, "%Y-%m-%d %H:%M:%S").timestamp())
    except ValueError:
        try:
            return int(datetime.strptime(s, "%d-%m-%Y %H:%M:%S").timestamp())
        except ValueError:
            return 0
